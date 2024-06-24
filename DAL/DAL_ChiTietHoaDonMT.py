from . import ConnectDB

def ShowAllCTHDMuonTra():
    conn = ConnectDB.connect_db()
    cs = conn.cursor()
    qr = "SELECT * FROM ctmuontra"
    cs.execute(qr)
    rs = cs.fetchall()
    conn.close()
    return rs

def SearchAllCTHDByMaMuonTra(maMuonTra):
    conn = ConnectDB.connect_db()
    qr = "SELECT * FROM ctmuontra WHERE mamuontra = %s"
    data = (maMuonTra,)
    cs = conn.cursor()
    cs.execute(qr, data)
    rs = cs.fetchall()
    conn.close()
    return rs


def DeleteCTHDByMaMuontra(maMuonTra):
    conn = ConnectDB.connect_db()
    qr = r"DELETE FROM ctmuontra WHERE mamuontra = %s"
    data = (maMuonTra,)
    cs = conn.cursor()
    cs.execute(qr, data)
    conn.commit()
    conn.close()
    return cs.rowcount

def DeleteCTHDByMaMuontraMaSach(maMuonTra, maSach):
    conn = ConnectDB.connect_db()
    qr = r"DELETE FROM ctmuontra WHERE mamuontra = %s AND masach = %s"
    data = (maMuonTra, maSach)
    cs = conn.cursor()
    cs.execute(qr, data)
    conn.commit()
    conn.close()
    return cs.rowcount

def UpdateCTHoaDonMuonTra(maMuonTra, maSach,tenSach, tenTG, Gia, soLuong, ngayMuon, ngayTraDuKien, ngayTraThucTe, trangThai):
    conn = ConnectDB.connect_db()
    qr = r"UPDATE ctmuontra SET maSach = %s, tenSach = %s, tenTG = %s, Gia = %s, soLuong = %s, ngayMuon = %s, ngayTraDuKien = %s, ngayTraThucTe = %s, trangThai = %s WHERE mamuontra = %s"
    data = (maSach, tenSach, tenTG, Gia, soLuong, ngayMuon, ngayTraDuKien, ngayTraThucTe, trangThai, maMuonTra)
    cs = conn.cursor()
    cs.execute(qr, data)
    conn.commit()
    conn.close()
    return cs.rowcount

def AddCTHoaDonMuonTra(maMuonTra, maSach,tenSach, tenTG, Gia, soLuong, ngayMuon, ngayTraDuKien, ngayTraThucTe, trangThai):
    conn = ConnectDB.connect_db()
    qr = r"INSERT INTO ctmuontra (mamuontra, masach, tenSach, tenTG, Gia, soLuong, ngayMuon, ngayTraDuKien, ngayTraThucTe, trangThai) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    data = (maMuonTra, maSach, tenSach, tenTG, Gia, soLuong, ngayMuon, ngayTraDuKien, ngayTraThucTe, trangThai)
    cs = conn.cursor()
    cs.execute(qr, data)
    conn.commit()
    conn.close()
    return cs.rowcount

def SearchCTHoaDonMuonTra(maMuonTra_TenSach):
    conn = ConnectDB.connect_db()
    qr = "SELECT * FROM ctmuontra WHERE mamuontra LIKE '%"+maMuonTra_TenSach+"%' OR tensach LIKE '%"+maMuonTra_TenSach+"%';"
    cs = conn.cursor()
    cs.execute(qr)
    rs = cs.fetchall()
    conn.close()
    return rs


import openpyxl

def XuatFileCTDHMuonTra(maMuonTra):
    conn = ConnectDB.connect_db()
    qr = "SELECT * FROM ctmuontra WHERE mamuontra = %s"
    data = (maMuonTra,)
    cs = conn.cursor()
    cs.execute(qr, data)
    rs = cs.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 8.5
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15

    ws.append(["mã hóa đơn", "Mã sách", "Tên sách", "Tên tác giả", "Giá", "Số lương", "Ngày mượn", "Ngày trả dự kiến", "Ngày trả thực tế", "Trạng thái"])
    sum = 0
    for row in range(rs.__len__()):
        sum += rs[row][4] * rs[row][5]
        ws.append(rs[row])
        
    ws.cell(rs.__len__()+4, 9, "Tổng tiền")
    ws.cell(rs.__len__()+4, 10, str(sum))
    wb.save('DanhSachChiTietMuonTra.xlsx')
